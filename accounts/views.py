from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .forms import UserRegistrationForm, UserLoginForm, UserProfileForm
from django.contrib.auth.forms import PasswordChangeForm
from django.core.cache import cache
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.db.models import Sum
from django.db.models import Sum, Count, Q, F 
import json
import base64
import os
import qrcode
from io import BytesIO

# ==================== HELPER FUNCTIONS ====================

def get_biometric_key(username):
    """Get cache key for biometric credentials"""
    return f"biometric_{username}"

def save_biometric_credential(username, credential_id):
    """Save biometric credential for user"""
    cache_key = get_biometric_key(username)
    credentials = cache.get(cache_key, [])
    if credential_id not in credentials:
        credentials.append(credential_id)
        cache.set(cache_key, credentials, 60*60*24*30)  # 30 days

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import login
from django.utils import timezone
import json
import secrets
import base64

@csrf_exempt
def biometric_login_challenge(request):
    """Generate a challenge for biometric login"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=400)
    
    try:
        data = json.loads(request.body)
        username = data.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        # Generate random challenge
        challenge = secrets.token_bytes(32)
        challenge_b64 = base64.b64encode(challenge).decode()
        
        # Store challenge in session for verification
        request.session['biometric_challenge'] = challenge_b64
        request.session['biometric_username'] = username
        
        return JsonResponse({'success': True, 'challenge': challenge_b64})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def biometric_verify_login(request):
    """Verify biometric login assertion"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=400)
    
    try:
        data = json.loads(request.body)
        username = data.get('username')
        credential_id = data.get('credential_id')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        # Find user
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)
        
        # Check if credential exists
        from accounts.models import BiometricCredential
        credential = BiometricCredential.objects.filter(
            user=user, 
            credential_id=credential_id, 
            is_active=True
        ).first()
        
        if not credential:
            return JsonResponse({'error': 'Credential not found'}, status=404)
        
        # Login the user
        login(request, user)
        
        # Update last used
        credential.last_used = timezone.now()
        credential.sign_count += 1
        credential.save(update_fields=['last_used', 'sign_count'])
        
        # Clear challenge from session
        if 'biometric_challenge' in request.session:
            del request.session['biometric_challenge']
        
        return JsonResponse({
            'success': True,
            'redirect_url': '/',
            'message': 'Login successful'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_biometric_credentials(request):
    """Get biometric credentials for a user (for login page)"""
    from accounts.models import BiometricCredential
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    username = request.GET.get('username')
    
    if not username:
        return JsonResponse({'credentials': []})
    
    try:
        user = User.objects.get(username=username)
        credentials = BiometricCredential.objects.filter(user=user, is_active=True)
        
        creds_list = []
        for cred in credentials:
            # Convert credential_id to base64 for WebAuthn
            cred_id_b64 = base64.b64encode(cred.credential_id.encode()).decode()
            creds_list.append({
                'credential_id': cred_id_b64,
                'id': cred.id
            })
        
        return JsonResponse({'credentials': creds_list})
    except User.DoesNotExist:
        return JsonResponse({'credentials': []})
def has_biometric_credential(username):
    """Check if user has biometric credentials"""
    return len(get_biometric_credentials(username)) > 0
from django.utils import timezone
def clear_biometric_credentials(username):
    """Clear biometric credentials for user"""
    cache_key = get_biometric_key(username)
    cache.delete(cache_key)


# ==================== AUTHENTICATION VIEWS ====================

def register_view(request):
    """User registration view"""
    if request.user.is_authenticated:
        return redirect('core:home')
    
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Welcome to Bantu Books Zambia, {user.get_full_name()}! 🎉')
            return redirect('core:home')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = UserRegistrationForm()
    
    return render(request, 'accounts/register.html', {'form': form})


from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.shortcuts import render, redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth import get_user_model
from accounts.models import BiometricCredential

User = get_user_model()

from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.http import JsonResponse
import json

User = get_user_model()

def has_biometric_credential(username):
    """Check if user has biometric credentials"""
    try:
        from accounts.models import BiometricCredential
        user = User.objects.get(username=username)
        return BiometricCredential.objects.filter(user=user, is_active=True).exists()
    except:
        return False


@csrf_protect
def login_view(request):
    """User login view with biometric support"""
    
    # Redirect if already logged in
    if request.user.is_authenticated:
        return redirect('core:home')
    
    # Handle regular login (not AJAX)
    if request.method == 'POST' and not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        login_input = request.POST.get('login', '').strip()
        password = request.POST.get('password', '')
        remember = request.POST.get('remember')
        
        # Validate input
        if not login_input or not password:
            messages.error(request, 'Please enter both username/email and password.')
            return render(request, 'accounts/login.html')
        
        # Try to find user by username or email
        user = None
        
        if '@' in login_input:
            # It's an email
            try:
                user_obj = User.objects.get(email=login_input)
                user = authenticate(request, username=user_obj.username, password=password)
            except User.DoesNotExist:
                user = None
        else:
            # It's a username
            user = authenticate(request, username=login_input, password=password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                
                # Set session expiry
                if not remember:
                    request.session.set_expiry(0)  # Session expires when browser closes
                else:
                    request.session.set_expiry(1209600)  # 2 weeks
                
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                
                # Redirect to next parameter or home
                next_url = request.GET.get('next')
                if next_url:
                    return redirect(next_url)
                return redirect('core:home')
            else:
                messages.error(request, 'Your account is disabled. Please contact support.')
        else:
            messages.error(request, 'Invalid username/email or password. Please try again.')
    
    return render(request, 'accounts/login.html')



@login_required
def logout_view(request):
    """User logout view"""
    logout(request)
    messages.info(request, 'You have been logged out successfully.')
    return redirect('core:home')


# ==================== PROFILE VIEWS ====================
@login_required
def profile_view(request):
    """User profile view and edit with statistics and activity"""
    from downloads.models import BookDownload, BookReadOnline, UserDownloadLimit
    from books.models import BookReview
    from django.db.models import Sum, Count
    from datetime import timedelta
    
    # Handle profile update
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('accounts:profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserProfileForm(instance=request.user)
    
    # Get user statistics
    total_downloads = BookDownload.objects.filter(user=request.user).count()
    total_reads = BookReadOnline.objects.filter(user=request.user).count()
    
    # Get download limit info
    try:
        download_limit = UserDownloadLimit.objects.get(user=request.user)
        daily_limit = download_limit.daily_limit
        remaining_today = max(0, (daily_limit or 0) - download_limit.downloads_today) if daily_limit else "Unlimited"
        subscription_tier = download_limit.subscription_tier
    except UserDownloadLimit.DoesNotExist:
        daily_limit = 20
        remaining_today = 20
        subscription_tier = 'basic'
    
    # Get user credits balance (add this field to User model if not exists)
    credits_balance = getattr(request.user, 'credits_balance', 0)
    
    # Check if user has active premium subscription
    is_premium_active = False
    if request.user.user_type == 'premium':
        if hasattr(request.user, 'subscription_expiry') and request.user.subscription_expiry:
            # Check if subscription is not expired
            if request.user.subscription_expiry > timezone.now():
                is_premium_active = True
            else:
                # Subscription expired - downgrade to basic
                request.user.user_type = 'basic'
                request.user.save()
                messages.warning(request, 'Your premium subscription has expired.')
        else:
            # No expiry date set, assume active
            is_premium_active = True
    elif request.user.user_type == 'lifetime':
        is_premium_active = True
    
    # Get recent activity
    recent_downloads = BookDownload.objects.filter(
        user=request.user
    ).select_related('book').order_by('-downloaded_at')[:5]
    
    recent_reads = BookReadOnline.objects.filter(
        user=request.user
    ).select_related('book').order_by('-last_activity')[:5]
    
    # Get user reviews
    user_reviews = BookReview.objects.filter(
        user=request.user
    ).select_related('book').order_by('-created_at')[:5]
    
    # Get monthly download statistics
    today = timezone.now().date()
    monthly_stats = []
    
    for i in range(6):
        month_date = today.replace(day=1) - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        
        month_count = BookDownload.objects.filter(
            user=request.user,
            downloaded_at__date__gte=month_start,
            downloaded_at__date__lte=month_end
        ).count()
        
        monthly_stats.insert(0, {
            'month': month_start.strftime('%B'),
            'count': month_count
        })
    
    # Get member since info
    member_since = request.user.date_joined
    member_days = (timezone.now() - member_since).days
    
    # Get total reading time
    total_reading_time = BookReadOnline.objects.filter(
        user=request.user
    ).aggregate(total=Sum('time_spent'))['total'] or 0
    
    # Format reading time
    if total_reading_time >= 3600:
        reading_hours = total_reading_time // 3600
        reading_minutes = (total_reading_time % 3600) // 60
        reading_time_display = f"{reading_hours}h {reading_minutes}m"
    elif total_reading_time >= 60:
        reading_minutes = total_reading_time // 60
        reading_time_display = f"{reading_minutes} minutes"
    else:
        reading_time_display = f"{total_reading_time} seconds"
    
    context = {
        'form': form,
        'user': request.user,
        'total_downloads': total_downloads,
        'total_reads': total_reads,
        'daily_limit': daily_limit if daily_limit else "Unlimited",
        'remaining_today': remaining_today,
        'subscription_tier': subscription_tier,
        'credits_balance': credits_balance,
        'is_premium_active': is_premium_active,
        'recent_downloads': recent_downloads,
        'recent_reads': recent_reads,
        'user_reviews': user_reviews,
        'monthly_stats': monthly_stats,
        'member_since': member_since,
        'member_days': member_days,
        'reading_time_display': reading_time_display,
        'total_reading_time_seconds': total_reading_time,
    }
    
    return render(request, 'accounts/profile.html', context)


@login_required
def change_password_view(request):
    """Change password view"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been changed successfully!')
            return redirect('accounts:profile')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'accounts/change_password.html', {'form': form})
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from accounts.models import User
from payments.models import Transaction, PaymentMethod
import logging
import json

logger = logging.getLogger(__name__)


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string
import logging
import json
from uuid import uuid4

logger = logging.getLogger(__name__)


@login_required
def upgrade_account_view(request):
    """View for upgrading user account"""
    
    subscription_plans = {
        'monthly': {
            'name': 'Monthly Premium (Chibombo)', 
            'price': 35, 
            'duration': 1,
            'savings': 0,
            'features': [
                'Unlimited downloads',
                'Access to all premium content',
                'No advertisements',
                'Priority support',
                'Download up to 500 books/month'
            ]
        },
        'quarterly': {
            'name': 'Quarterly Premium (Chibombo)', 
            'price': 90, 
            'duration': 3,
            'savings': 15,
            'features': [
                'Unlimited downloads',
                'Access to all premium content',
                'No advertisements',
                'Priority support',
                'Save 15% compared to monthly',
                'Download up to 2000 books/month'
            ]
        },
        'yearly': {
            'name': 'Yearly Premium (Chibombo)', 
            'price': 350, 
            'duration': 12,
            'savings': 70,
            'features': [
                'Unlimited downloads',
                'Access to all premium content',
                'No advertisements',
                'Priority support',
                'Save 20% compared to monthly',
                'Early access to new content',
                'Download up to 10000 books/year'
            ]
        },
        'lifetime': {
            'name': 'Lifetime Access (Chifumu)', 
            'price': 800, 
            'duration': None,
            'savings': 0,
            'features': [
                'Unlimited downloads forever',
                'Access to all premium content',
                'No advertisements',
                'Priority support',
                'Free print coupons annually',
                'Name in supporters list',
                'Never expires'
            ]
        },
    }
    
    # Calculate savings for display
    monthly_base = 35
    for key, plan in subscription_plans.items():
        if plan.get('duration') and plan['duration'] > 1:
            monthly_cost = plan['price'] / plan['duration']
            plan['calculated_savings'] = round((monthly_base - monthly_cost) * plan['duration'], 2)
        else:
            plan['calculated_savings'] = 0
    
    context = {
        'plans': subscription_plans,
        'current_user_type': request.user.user_type,
        'user': request.user,
    }
    
    return render(request, 'accounts/upgrade.html', context)


@login_required
def process_upgrade_payment(request):
    """Process upgrade payment via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    
    try:
        data = json.loads(request.body)
        plan_key = data.get('plan')
        payment_method = data.get('payment_method', 'auto')
        
        # Define plans
        plans = {
            'monthly': {'name': 'Monthly Premium', 'price': 35, 'duration_days': 30},
            'quarterly': {'name': 'Quarterly Premium', 'price': 90, 'duration_days': 90},
            'yearly': {'name': 'Yearly Premium', 'price': 350, 'duration_days': 365},
            'lifetime': {'name': 'Lifetime Access', 'price': 800, 'duration_days': None},
        }
        
        if plan_key not in plans:
            return JsonResponse({'success': False, 'error': 'Invalid plan selected'}, status=400)
        
        plan = plans[plan_key]
        amount = plan['price']
        
        # Check if user already has an active premium subscription
        if request.user.user_type in ['premium', 'lifetime']:
            return JsonResponse({
                'success': False, 
                'error': 'You already have an active premium subscription.'
            }, status=400)
        
        # Get payment details
        phone_number = data.get('phone_number', '')
        bank_details = data.get('bank_details', {})
        
        # Create transaction record
        from downloads.models import Transaction
        
        transaction = Transaction.objects.create(
            user=request.user,
            transaction_id=f"UPG_{int(timezone.now().timestamp())}_{request.user.id}",
            amount=amount,
            plan_type=plan_key,
            payment_method=payment_method,
            status='completed',  # Set to completed directly for demo
            payment_details={
                'method': payment_method,
                'plan': plan_key,
                'phone_number': phone_number,
                'bank_details': bank_details
            },
            payment_confirmed_at=timezone.now()
        )
        
        # Upgrade the user immediately
        success, message = upgrade_user_subscription(request.user, plan_key, transaction)
        
        if success:
            # Send email notification
            send_upgrade_notification(request.user, plan_key, amount)
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully upgraded to {plan["name"]}!',
                'redirect_url': reverse('accounts:subscription_dashboard')
            })
        else:
            return JsonResponse({'success': False, 'error': message}, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Payment processing error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
def upgrade_user_subscription(user, plan_type, transaction=None):
    """Upgrade user account based on selected plan"""
    try:
        from downloads.models import UserDownloadLimit
        
        plans = {
            'monthly': {'duration_days': 30, 'user_type': 'premium', 'price': 35},
            'quarterly': {'duration_days': 90, 'user_type': 'premium', 'price': 90},
            'yearly': {'duration_days': 365, 'user_type': 'premium', 'price': 350},
            'lifetime': {'duration_days': None, 'user_type': 'lifetime', 'price': 800}
        }
        
        if plan_type not in plans:
            return False, "Invalid plan type"
        
        plan = plans[plan_type]
        
        # Update user type
        old_type = user.user_type
        user.user_type = plan['user_type']
        
        # Set subscription expiry
        if plan['duration_days']:
            user.subscription_expiry = timezone.now() + timedelta(days=plan['duration_days'])
        else:
            user.subscription_expiry = None  # Never expires
        
        user.save()
        
        # Update or create download limits
        download_limit, created = UserDownloadLimit.objects.get_or_create(user=user)
        
        if plan_type in ['monthly', 'quarterly', 'yearly']:
            download_limit.subscription_tier = 'premium'
            download_limit.daily_limit = None  # Unlimited
            download_limit.monthly_limit = None  # Unlimited
        elif plan_type == 'lifetime':
            download_limit.subscription_tier = 'lifetime'
            download_limit.daily_limit = None
            download_limit.monthly_limit = None
        download_limit.save()
        
        # Update transaction if provided
        if transaction:
            transaction.status = 'completed'
            transaction.payment_confirmed_at = timezone.now()
            transaction.save()
        
        logger.info(f"User {user.email} upgraded from {old_type} to {plan['user_type']} with {plan_type} plan")
        return True, "Upgrade successful"
        
    except Exception as e:
        logger.error(f"Error upgrading user: {e}")
        return False, str(e)


def send_upgrade_notification(user, plan_type, amount):
    """Send email notification for successful upgrade"""
    try:
        subject = f"Account Upgraded to Premium - Bantu Books Zambia"
        html_message = render_to_string('emails/upgrade_confirmation.html', {
            'user': user,
            'plan_type': plan_type,
            'amount': amount,
            'upgrade_date': timezone.now(),
            'expiry_date': user.subscription_expiry,
        })
        plain_message = f"Dear {user.username},\n\nYour account has been upgraded to {plan_type.upper()} plan. Thank you for your support!\n\nRegards,\nBantu Books Zambia Team"
        
        send_mail(
            subject,
            plain_message,
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            html_message=html_message,
            fail_silently=True,
        )
    except Exception as e:
        logger.error(f"Failed to send upgrade email: {e}")
@login_required
def subscription_dashboard(request):
    """Display user's subscription information and usage statistics"""
    from downloads.models import BookDownload, UserDownloadLimit
    from django.db.models import Sum
    from datetime import timedelta
    from django.utils import timezone
    
    # ============ 1. USER TYPE & LIMITS SETUP ============
    user_type_map = {
        'basic': ('basic', 20, 500),
        'premium': ('premium', None, None),
        'lifetime': ('lifetime', None, None)
    }
    
    current_tier = user_type_map.get(request.user.user_type, ('basic', 20, 500))[0]
    
    # Get or create user download limit
    download_limit, created = UserDownloadLimit.objects.get_or_create(
        user=request.user,
        defaults={
            'subscription_tier': current_tier,
            'daily_limit': 20 if current_tier == 'basic' else None,
            'monthly_limit': 500 if current_tier == 'basic' else None
        }
    )
    
    # Update download limit if tier changed
    tier_info = user_type_map.get(request.user.user_type, ('basic', 20, 500))
    if download_limit.subscription_tier != tier_info[0]:
        download_limit.subscription_tier = tier_info[0]
        download_limit.daily_limit = tier_info[1]
        download_limit.monthly_limit = tier_info[2]
        download_limit.save()
    
    # ============ 2. DATE RANGE CALCULATIONS ============
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)
    last_week_start = start_of_week - timedelta(days=7)
    last_month_start = start_of_month - timedelta(days=30)
    
    # ============ 3. DOWNLOAD STATISTICS ============
    # Today
    today_downloads = BookDownload.objects.filter(
        user=request.user, downloaded_at__date=today
    ).count()
    
    # This week
    week_downloads = BookDownload.objects.filter(
        user=request.user, downloaded_at__date__gte=start_of_week
    ).count()
    
    # This month
    month_downloads = BookDownload.objects.filter(
        user=request.user, downloaded_at__date__gte=start_of_month
    ).count()
    
    # All time
    total_downloads = BookDownload.objects.filter(user=request.user).count()
    
    # Last week (for trend)
    last_week_downloads = BookDownload.objects.filter(
        user=request.user,
        downloaded_at__date__gte=last_week_start,
        downloaded_at__date__lt=start_of_week
    ).count()
    
    # Last month (for trend)
    last_month_downloads = BookDownload.objects.filter(
        user=request.user,
        downloaded_at__date__gte=last_month_start,
        downloaded_at__date__lt=start_of_month
    ).count()
    
    # ============ 4. WEEKLY DAILY BREAKDOWN ============
    week_daily_data = []
    week_max = 0
    best_day_count = 0
    best_day_name = ""
    
    for i in range(7):
        day_date = start_of_week + timedelta(days=i)
        day_name = day_date.strftime('%A')
        day_short = day_date.strftime('%a')
        
        day_count = BookDownload.objects.filter(
            user=request.user,
            downloaded_at__date=day_date
        ).count()
        
        week_daily_data.append({
            'name': day_name,
            'short_name': day_short,
            'count': day_count,
            'date': day_date
        })
        
        if day_count > week_max:
            week_max = day_count
        
        if day_count > best_day_count:
            best_day_count = day_count
            best_day_name = day_name
    
    week_avg = week_downloads / 7 if week_downloads > 0 else 0
    
    # ============ 5. TREND CALCULATIONS ============
    week_trend = 0
    if last_week_downloads > 0:
        week_trend = round(((week_downloads - last_week_downloads) / last_week_downloads) * 100, 1)
    elif week_downloads > 0:
        week_trend = 100
    
    month_trend = 0
    if last_month_downloads > 0:
        month_trend = round(((month_downloads - last_month_downloads) / last_month_downloads) * 100, 1)
    elif month_downloads > 0:
        month_trend = 100
    
    # ============ 6. UPDATE DOWNLOAD COUNTS ============
    download_limit.downloads_today = today_downloads
    download_limit.downloads_this_month = month_downloads
    download_limit.downloads_total = total_downloads
    download_limit.save()
    
    # ============ 7. RECENT DOWNLOADS ============
    recent_downloads = BookDownload.objects.filter(
        user=request.user
    ).select_related('book').order_by('-downloaded_at')[:10]
    
    # ============ 8. REMAINING COUNTS ============
    if download_limit.subscription_tier in ['premium', 'lifetime']:
        remaining_today = "Unlimited"
        remaining_month = "Unlimited"
        remaining_today_display = "∞"
    else:
        remaining_today = max(0, (download_limit.daily_limit or 0) - today_downloads)
        remaining_month = max(0, (download_limit.monthly_limit or 0) - month_downloads)
        remaining_today_display = remaining_today
    
    # ============ 9. SUBSCRIPTION STATUS & CREDITS ============
    is_subscription_active = False
    subscription_expiry = None
    credits_balance = getattr(request.user, 'credits_balance', 0)
    
    if request.user.user_type == 'premium':
        if hasattr(request.user, 'subscription_expiry') and request.user.subscription_expiry:
            if request.user.subscription_expiry > timezone.now():
                is_subscription_active = True
                subscription_expiry = request.user.subscription_expiry
            else:
                # Subscription expired - downgrade to basic
                request.user.user_type = 'basic'
                request.user.save()
                from django.contrib import messages
                messages.warning(request, 'Your premium subscription has expired. Please renew to continue enjoying premium benefits.')
        else:
            is_subscription_active = True
    elif request.user.user_type == 'lifetime':
        is_subscription_active = True
    
    # ============ 10. PLAN INFORMATION ============
    plan_info = {
        'basic': {
            'name': 'Basic (Mumbi)',
            'price': 'Free',
            'download_limit': '20 downloads/day',
            'features': [
                '20 downloads per day',
                'Access to free books only',
                'Basic support',
                'Standard download speed'
            ]
        },
        'premium': {
            'name': 'Premium (Chibombo)',
            'price': 'K35/month',
            'download_limit': 'Unlimited',
            'features': [
                'Unlimited downloads',
                'Access to ALL books',
                'No advertisements',
                'Priority support',
                'Early access to new books'
            ]
        },
        'lifetime': {
            'name': 'Lifetime (Chifumu)',
            'price': 'K800 one-time',
            'download_limit': 'Unlimited',
            'features': [
                'Unlimited downloads forever',
                'Access to ALL books',
                'No advertisements',
                'Priority support',
                'Early access to new books',
                'Name in supporters list',
                'Never expires'
            ]
        }
    }
    
    current_plan = plan_info.get(current_tier, plan_info['basic'])
    
    # ============ 11. LIMITS FOR DISPLAY ============
    daily_limit = download_limit.daily_limit if download_limit.daily_limit is not None else "Unlimited"
    monthly_limit = download_limit.monthly_limit if download_limit.monthly_limit is not None else "Unlimited"
    
    # ============ 12. MONTHLY STATS FOR CHART ============
    monthly_stats = []
    for i in range(6):
        month_date = today.replace(day=1) - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        
        month_count = BookDownload.objects.filter(
            user=request.user,
            downloaded_at__date__gte=month_start,
            downloaded_at__date__lte=month_end
        ).count()
        
        monthly_stats.insert(0, {
            'month': month_start.strftime('%b'),
            'count': month_count
        })
    
    # ============ 13. BUILD CONTEXT ============
    context = {
        # Plan info
        'current_plan': current_plan,
        'current_tier': current_tier,
        'plan_info': plan_info,
        
        # Download counts
        'total_downloads': total_downloads,
        'today_downloads': today_downloads,
        'week_downloads': week_downloads,
        'month_downloads': month_downloads,
        
        # Remaining counts
        'remaining_today': remaining_today,
        'remaining_month': remaining_month,
        'remaining_today_display': remaining_today_display,
        'daily_limit': daily_limit,
        'monthly_limit': monthly_limit,
        
        # Trends
        'week_trend': week_trend,
        'month_trend': month_trend,
        'last_week_downloads': last_week_downloads,
        'last_month_downloads': last_month_downloads,
        
        # Weekly breakdown
        'week_daily_data': week_daily_data,
        'week_max': week_max,
        'week_avg': round(week_avg, 1),
        'best_day_count': best_day_count,
        'best_day_name': best_day_name,
        
        # Recent activity
        'recent_downloads': recent_downloads,
        
        # Subscription status
        'is_subscription_active': is_subscription_active,
        'subscription_expiry': subscription_expiry,
        'credits_balance': credits_balance,
        
        # Date info
        'today': today,
        'week_start_date': start_of_week,
        'week_end_date': start_of_week + timedelta(days=6),
        'monthly_stats': monthly_stats,
    }
    
    return render(request, 'accounts/subscription_dashboard.html', context)
@login_required
def get_payment_status(request):
    """Get current payment status for user"""
    from downloads.models import Transaction
    
    recent_transactions = Transaction.objects.filter(
        user=request.user
    ).order_by('-created_at')[:5]
    
    transactions_data = []
    for t in recent_transactions:
        transactions_data.append({
            'id': t.transaction_id,
            'amount': str(t.amount),
            'plan_type': t.plan_type,
            'status': t.status,
            'created_at': t.created_at.isoformat(),
            'completed_at': t.payment_confirmed_at.isoformat() if t.payment_confirmed_at else None
        })
    
    return JsonResponse({
        'has_transactions': len(transactions_data) > 0,
        'transactions': transactions_data,
        'current_user_type': request.user.user_type,
        'subscription_expiry': request.user.subscription_expiry.isoformat() if request.user.subscription_expiry else None
    })
# ==================== HISTORY VIEWS ====================

@login_required
def download_history_view(request):
    """View user's download history"""
    try:
        from downloads.models import BookDownload
        downloads = BookDownload.objects.filter(user=request.user).select_related('book')[:50]
    except ImportError:
        downloads = []
    
    return render(request, 'accounts/download_history.html', {
        'downloads': downloads,
    })


@login_required
def reading_history_view(request):
    """View user's reading history"""
    try:
        from downloads.models import BookReadOnline
        reads = BookReadOnline.objects.filter(user=request.user).select_related('book')[:50]
    except ImportError:
        reads = []
    
    return render(request, 'accounts/reading_history.html', {
        'reads': reads,
    })


# ==================== BIOMETRIC / FINGERPRINT VIEWS ====================

import json
import hashlib
import secrets
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import login
from django.core.cache import cache
from django.conf import settings
import logging

logger = logging.getLogger(__name__)

# In-memory storage for demo (replace with database in production)
# For production, create a BiometricCredential model
_biometric_storage = {}


def has_biometric_credential(username):
    """Check if user has a biometric credential"""
    return username in _biometric_storage and len(_biometric_storage[username]) > 0





def save_biometric_credential(username, credential_data):
    """Save biometric credential for user"""
    if username not in _biometric_storage:
        _biometric_storage[username] = []
    
    # Check if credential already exists
    credential_id = credential_data.get('id') or credential_data.get('credential_id')
    existing = [c for c in _biometric_storage[username] if c.get('id') == credential_id]
    
    if not existing:
        _biometric_storage[username].append({
            'id': credential_id,
            'data': credential_data,
            'created_at': timezone.now().isoformat(),
            'device_name': credential_data.get('device_name', 'Unknown Device')
        })
        return True
    return False


def clear_biometric_credentials(username):
    """Clear all biometric credentials for user"""
    if username in _biometric_storage:
        del _biometric_storage[username]
        return True
    return False


@login_required
def security_settings(request):
    """Security settings page for biometric authentication"""
    has_biometric = has_biometric_credential(request.user.username)
    credentials = get_biometric_credentials(request.user.username)
    
    context = {
        'user': request.user,
        'has_biometric': has_biometric,
        'credentials': credentials,
        'webauthn_supported': check_webauthn_support(),
    }
    return render(request, 'accounts/security_settings.html', context)


def check_webauthn_support():
    """Check if WebAuthn is supported (client-side check in template)"""
    # This will be handled in the template
    return True

@login_required
def register_biometric(request):
    """Register biometric credential using WebAuthn"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
    try:
        import json
        import base64
        from accounts.models import BiometricCredential
        from django.utils import timezone
        
        data = json.loads(request.body)
        
        # Get credential data
        credential_id = data.get('id')
        
        if not credential_id:
            return JsonResponse({'error': 'No credential ID provided'}, status=400)
        
        # Delete any existing credentials first (re-registration)
        BiometricCredential.objects.filter(user=request.user).delete()
        
        # Get device name from user agent
        device_name = data.get('device_name', 'WebAuthn Device')
        
        # Convert rawId from base64 to string
        raw_id = data.get('rawId', '')
        
        # Save new credential
        credential = BiometricCredential.objects.create(
            user=request.user,
            credential_id=credential_id,
            public_key=json.dumps(data.get('response', {})),
            device_name=device_name,
            is_active=True,
            created_at=timezone.now()
        )
        
        print(f"Biometric registered for user: {request.user.username} - Credential ID: {credential.id}")
        
        return JsonResponse({
            'success': True,
            'message': 'Biometric registered successfully',
            'credential_id': credential_id
        })
        
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {e}")
        return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
    except Exception as e:
        print(f"Registration error: {e}")
        return JsonResponse({'error': str(e)}, status=500)
@login_required
def remove_biometric(request):
    """Remove biometric authentication"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body) if request.body else {}
            credential_id = data.get('credential_id')
            
            if credential_id and credential_id in _biometric_storage.get(request.user.username, []):
                _biometric_storage[request.user.username] = [
                    c for c in _biometric_storage[request.user.username] 
                    if c.get('id') != credential_id and c.get('credential_id') != credential_id
                ]
                messages.success(request, 'Biometric credential removed successfully.')
            else:
                clear_biometric_credentials(request.user.username)
                messages.success(request, 'Biometric authentication has been removed.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Biometric removed successfully'})
            
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': str(e)}, status=500)
            messages.error(request, f'Error removing biometric: {str(e)}')
        
        return redirect('accounts:security_settings')
    
    return redirect('accounts:security_settings')






def generate_webauthn_challenge(request):
    """Generate a WebAuthn challenge for registration"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
    try:
        # Generate a random challenge
        challenge = secrets.token_bytes(32)
        challenge_b64 = base64.b64encode(challenge).decode('utf-8')
        
        # Store challenge in session for verification
        request.session['webauthn_challenge'] = challenge_b64
        
        return JsonResponse({
            'success': True,
            'challenge': challenge_b64,
            'rp_id': request.get_host().split(':')[0],  # Remove port if present
            'rp_name': 'Bantu Books Zambia'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def verify_webauthn_registration(request):
    """Verify WebAuthn registration"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
    try:
        data = json.loads(request.body)
        credential_id = data.get('id')
        
        if not credential_id:
            return JsonResponse({'error': 'Missing credential ID'}, status=400)
        
        # Store credential (you would normally store the public key)
        save_biometric_credential(request.user.username, {
            'id': credential_id,
            'credential_id': credential_id,
            'raw_id': data.get('rawId'),
            'response': data.get('response'),
            'type': data.get('type')
        })
        
        # Clear challenge from session
        if 'webauthn_challenge' in request.session:
            del request.session['webauthn_challenge']
        
        return JsonResponse({
            'success': True,
            'message': 'WebAuthn registered successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Add timezone import at the top
from django.utils import timezone
import base64
# ==================== TWO-FACTOR AUTHENTICATION VIEWS ====================

def enable_2fa(request):
    """Enable two-factor authentication"""
    if not request.user.is_authenticated:
        return redirect('accounts:login')
    
    if request.method == 'POST':
        token = request.POST.get('token')
        try:
            from django_otp.plugins.otp_totp.models import TOTPDevice
            from django_otp.plugins.otp_static.models import StaticDevice
            
            device, created = TOTPDevice.objects.get_or_create(user=request.user, name='default')
            if device.verify_token(token):
                device.confirmed = True
                device.save()
                
                # Generate backup codes
                static_device, _ = StaticDevice.objects.get_or_create(user=request.user, name='backup')
                backup_codes = []
                for i in range(10):
                    code = static_device.generate_token()
                    backup_codes.append(code)
                request.session['backup_codes'] = backup_codes
                
                messages.success(request, '2FA enabled successfully!')
                return redirect('accounts:profile')
            else:
                messages.error(request, 'Invalid code. Please try again.')
        except ImportError:
            messages.error(request, '2FA feature is not available.')
    
    return redirect('accounts:profile')


def disable_2fa(request):
    """Disable two-factor authentication"""
    if not request.user.is_authenticated:
        return redirect('accounts:login')
    
    if request.method == 'POST':
        try:
            from django_otp.plugins.otp_totp.models import TOTPDevice
            from django_otp.plugins.otp_static.models import StaticDevice
            
            TOTPDevice.objects.filter(user=request.user).delete()
            StaticDevice.objects.filter(user=request.user).delete()
            messages.success(request, '2FA disabled successfully.')
        except ImportError:
            messages.error(request, '2FA feature is not available.')
    
    return redirect('accounts:profile')


def qrcode_view(request):
    """Generate QR code for 2FA setup"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    try:
        from django_otp.plugins.otp_totp.models import TOTPDevice
        
        device, created = TOTPDevice.objects.get_or_create(user=request.user, name='default')
        provisioning_uri = device.config_url
        
        img = qrcode.make(provisioning_uri)
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        img_str = base64.b64encode(buffer.getvalue()).decode()
        
        return JsonResponse({'qr_code': f'data:image/png;base64,{img_str}'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def generate_backup_codes(request):
    """Generate new backup codes"""
    if not request.user.is_authenticated:
        return redirect('accounts:login')
    
    if request.method == 'POST':
        try:
            from django_otp.plugins.otp_static.models import StaticDevice
            
            StaticDevice.objects.filter(user=request.user).delete()
            static_device = StaticDevice.objects.create(user=request.user, name='backup')
            
            backup_codes = []
            for i in range(10):
                code = static_device.generate_token()
                backup_codes.append(code)
            
            # Send backup codes via email
            email_body = "Your backup codes for Bantu Books Zambia:\n\n"
            for code in backup_codes:
                email_body += f"{code}\n"
            email_body += "\nKeep these codes safe. Each code can be used only once."
            
            send_mail(
                'Your Backup Codes',
                email_body,
                settings.DEFAULT_FROM_EMAIL,
                [request.user.email],
                fail_silently=True,
            )
            
            messages.success(request, 'New backup codes have been generated and sent to your email.')
        except ImportError:
            messages.error(request, '2FA feature is not available.')
    
    return redirect('accounts:profile')
# Add these missing functions to your accounts/views.py

@login_required
def verify_payment(request, transaction_id):
    """Verify payment status (AJAX)"""
    try:
        from downloads.models import Transaction
        
        transaction = Transaction.objects.get(transaction_id=transaction_id, user=request.user)
        
        return JsonResponse({
            'success': True,
            'status': transaction.status,
            'amount': str(transaction.amount),
            'plan_type': transaction.plan_type,
            'created_at': transaction.created_at.isoformat(),
            'completed_at': transaction.payment_confirmed_at.isoformat() if transaction.payment_confirmed_at else None,
            'message': f'Payment status: {transaction.get_status_display()}'
        })
    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def download_history_view(request):
    """View user's download history"""
    from downloads.models import BookDownload
    from django.core.paginator import Paginator
    
    downloads = BookDownload.objects.filter(
        user=request.user
    ).select_related('book').order_by('-downloaded_at')
    
    paginator = Paginator(downloads, 20)
    page = request.GET.get('page', 1)
    downloads_page = paginator.get_page(page)
    
    context = {
        'downloads': downloads_page,
        'total_downloads': downloads.count(),
    }
    return render(request, 'accounts/download_history.html', context)


@login_required
def reading_history_view(request):
    """View user's reading history"""
    from downloads.models import BookReadOnline
    from django.core.paginator import Paginator
    
    readings = BookReadOnline.objects.filter(
        user=request.user
    ).select_related('book').order_by('-last_activity')
    
    paginator = Paginator(readings, 20)
    page = request.GET.get('page', 1)
    readings_page = paginator.get_page(page)
    
    context = {
        'readings': readings_page,
        'total_readings': readings.count(),
    }
    return render(request, 'accounts/reading_history.html', context)


@login_required
def security_settings(request):
    """View for security settings (2FA, biometric, etc.)"""
    from accounts.models import BiometricDevice, TwoFactorAuth, LoginHistory
    from django.db.models import Count
    from datetime import timedelta
    
    # Get biometric status
    biometric_devices = BiometricDevice.objects.filter(user=request.user, is_active=True)
    has_biometric = biometric_devices.exists()
    
    # Get 2FA status
    two_factor_auth = TwoFactorAuth.objects.filter(user=request.user).first()
    has_2fa = two_factor_auth and two_factor_auth.is_enabled
    
    # Get recent login history
    recent_logins = LoginHistory.objects.filter(
        user=request.user
    ).order_by('-login_time')[:5]
    
    # Get login statistics
    last_7_days = timezone.now() - timedelta(days=7)
    logins_last_7_days = LoginHistory.objects.filter(
        user=request.user,
        login_time__gte=last_7_days,
        login_success=True
    ).count()
    
    # Get failed login attempts in last 24 hours
    last_24_hours = timezone.now() - timedelta(hours=24)
    failed_logins = LoginHistory.objects.filter(
        user=request.user,
        login_time__gte=last_24_hours,
        login_success=False
    ).count()
    
    # Get last password change
    last_password_change = request.user.last_login if request.user.last_login else None
    
    # Check if user has recovery email set
    has_recovery_email = bool(request.user.email)
    
    # Get active sessions (simplified - in production, use django-user-sessions)
    active_sessions_count = 1  # Current session
    
    # Security score calculation (0-100)
    security_score = 0
    security_items = []
    
    # Check various security features
    if has_2fa:
        security_score += 30
        security_items.append({'name': 'Two-Factor Authentication', 'status': 'enabled', 'icon': 'fa-mobile-alt'})
    else:
        security_items.append({'name': 'Two-Factor Authentication', 'status': 'disabled', 'icon': 'fa-mobile-alt', 'action': 'enable'})
    
    if has_biometric:
        security_score += 25
        security_items.append({'name': 'Biometric Login', 'status': 'enabled', 'icon': 'fa-fingerprint'})
    else:
        security_items.append({'name': 'Biometric Login', 'status': 'disabled', 'icon': 'fa-fingerprint', 'action': 'setup'})
    
    # Check password strength (simplified)
    password_length = len(request.user.password) if request.user.password else 0
    if password_length > 0:
        security_score += 20
        security_items.append({'name': 'Password Set', 'status': 'enabled', 'icon': 'fa-key'})
    
    if has_recovery_email:
        security_score += 15
        security_items.append({'name': 'Recovery Email', 'status': 'enabled', 'icon': 'fa-envelope'})
    else:
        security_items.append({'name': 'Recovery Email', 'status': 'disabled', 'icon': 'fa-envelope', 'action': 'add'})
    
    if failed_logins == 0:
        security_score += 10
        security_items.append({'name': 'No Failed Login Attempts', 'status': 'good', 'icon': 'fa-shield-alt'})
    
    # Get device information
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    device_type = 'Unknown'
    browser = 'Unknown'
    os = 'Unknown'
    
    user_agent_lower = user_agent.lower()
    if 'mobile' in user_agent_lower or 'android' in user_agent_lower or 'iphone' in user_agent_lower:
        device_type = 'Mobile'
    elif 'tablet' in user_agent_lower or 'ipad' in user_agent_lower:
        device_type = 'Tablet'
    else:
        device_type = 'Desktop'
    
    if 'chrome' in user_agent_lower:
        browser = 'Chrome'
    elif 'firefox' in user_agent_lower:
        browser = 'Firefox'
    elif 'safari' in user_agent_lower:
        browser = 'Safari'
    elif 'edge' in user_agent_lower:
        browser = 'Edge'
    
    if 'windows' in user_agent_lower:
        os = 'Windows'
    elif 'mac' in user_agent_lower:
        os = 'macOS'
    elif 'linux' in user_agent_lower:
        os = 'Linux'
    elif 'android' in user_agent_lower:
        os = 'Android'
    elif 'ios' in user_agent_lower or 'iphone' in user_agent_lower or 'ipad' in user_agent_lower:
        os = 'iOS'
    
    context = {
        # Biometric
        'has_biometric': has_biometric,
        'biometric_devices': biometric_devices,
        'biometric_count': biometric_devices.count(),
        
        # 2FA
        'has_2fa': has_2fa,
        'two_factor_auth': two_factor_auth,
        
        # Login History
        'recent_logins': recent_logins,
        'logins_last_7_days': logins_last_7_days,
        'failed_logins': failed_logins,
        
        # Security Info
        'last_password_change': last_password_change,
        'has_recovery_email': has_recovery_email,
        'active_sessions_count': active_sessions_count,
        
        # Security Score
        'security_score': security_score,
        'security_items': security_items,
        
        # Device Info
        'current_device': {
            'type': device_type,
            'browser': browser,
            'os': os,
            'ip': request.META.get('REMOTE_ADDR'),
        },
        
        # User
        'user': request.user,
    }
    
    return render(request, 'accounts/security_settings.html', context)


@login_required
def enable_2fa(request):
    """Enable two-factor authentication"""
    import pyotp
    import qrcode
    from io import BytesIO
    import base64
    from accounts.models import TwoFactorAuth
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            otp_code = data.get('otp_code')
            secret_key = request.session.get('2fa_secret_key')
            
            if not secret_key:
                return JsonResponse({'success': False, 'error': 'Session expired. Please try again.'}, status=400)
            
            totp = pyotp.TOTP(secret_key)
            if totp.verify(otp_code):
                twofa, created = TwoFactorAuth.objects.get_or_create(user=request.user)
                twofa.is_enabled = True
                twofa.secret_key = secret_key
                twofa.save()
                
                # Generate backup codes
                backup_codes = twofa.generate_backup_codes()
                
                request.session.pop('2fa_secret_key', None)
                
                return JsonResponse({
                    'success': True,
                    'message': '2FA enabled successfully',
                    'backup_codes': backup_codes
                })
            else:
                return JsonResponse({'success': False, 'error': 'Invalid OTP code'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    # GET request - generate secret key
    import pyotp
    secret_key = pyotp.random_base32()
    request.session['2fa_secret_key'] = secret_key
    
    totp = pyotp.TOTP(secret_key)
    provisioning_uri = totp.provisioning_uri(request.user.email, issuer_name="Bantu Books Zambia")
    
    # Generate QR code
    qr = qrcode.QRCode(box_size=10, border=4)
    qr.add_data(provisioning_uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    return JsonResponse({
        'success': True,
        'secret_key': secret_key,
        'qr_code': qr_code_base64,
        'provisioning_uri': provisioning_uri
    })


@login_required
def disable_2fa(request):
    """Disable two-factor authentication"""
    if request.method == 'POST':
        try:
            from accounts.models import TwoFactorAuth
            
            twofa = TwoFactorAuth.objects.filter(user=request.user).first()
            if twofa:
                twofa.is_enabled = False
                twofa.secret_key = ''
                twofa.backup_codes = []
                twofa.save()
            
            return JsonResponse({'success': True, 'message': '2FA disabled successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@login_required
def generate_backup_codes(request):
    """Generate new backup codes for 2FA"""
    if request.method == 'POST':
        try:
            from accounts.models import TwoFactorAuth
            
            twofa = TwoFactorAuth.objects.get(user=request.user, is_enabled=True)
            backup_codes = twofa.generate_backup_codes()
            
            return JsonResponse({
                'success': True,
                'backup_codes': backup_codes,
                'message': 'New backup codes generated successfully'
            })
        except TwoFactorAuth.DoesNotExist:
            return JsonResponse({'success': False, 'error': '2FA not enabled'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@login_required
def qrcode_view(request):
    """Generate QR code for 2FA setup"""
    try:
        import pyotp
        import qrcode
        from io import BytesIO
        import base64
        
        secret_key = request.session.get('2fa_secret_key')
        if not secret_key:
            return JsonResponse({'success': False, 'error': 'No 2FA session found'}, status=400)
        
        totp = pyotp.TOTP(secret_key)
        provisioning_uri = totp.provisioning_uri(request.user.email, issuer_name="Bantu Books Zambia")
        
        qr = qrcode.QRCode(box_size=10, border=4)
        qr.add_data(provisioning_uri)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        return JsonResponse({
            'success': True,
            'qr_code': qr_code_base64,
            'secret_key': secret_key
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def forgot_password(request):
    """Handle forgot password request"""
    from django.contrib.auth.forms import PasswordResetForm
    from django.contrib.auth.models import User as AuthUser
    
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                user = AuthUser.objects.get(email=email)
                # Send password reset email
                form.save(
                    request=request,
                    use_https=request.is_secure(),
                    email_template_name='accounts/password_reset_email.html',
                )
                messages.success(request, 'Password reset link has been sent to your email.')
                return redirect('accounts:login')
            except AuthUser.DoesNotExist:
                messages.error(request, 'No user found with this email address.')
        else:
            messages.error(request, 'Please enter a valid email address.')
    else:
        form = PasswordResetForm()
    
    return render(request, 'accounts/forgot_password.html', {'form': form})


def reset_password(request, uidb64, token):
    """Handle password reset confirmation"""
    from django.contrib.auth.forms import SetPasswordForm
    from django.utils.encoding import force_str
    from django.utils.http import urlsafe_base64_decode
    from django.contrib.auth.tokens import default_token_generator
    from django.contrib.auth.models import User as AuthUser
    
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = AuthUser.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, AuthUser.DoesNotExist):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Your password has been reset. You can now log in.')
                return redirect('accounts:login')
            else:
                messages.error(request, 'Please correct the errors below.')
        else:
            form = SetPasswordForm(user)
        
        return render(request, 'accounts/reset_password.html', {'form': form})
    else:
        messages.error(request, 'The password reset link is invalid or has expired.')
        return redirect('accounts:forgot_password')


@login_required
def remove_biometric(request):
    """Remove biometric device for user"""
    if request.method == 'POST':
        try:
            from accounts.models import BiometricDevice
            
            BiometricDevice.objects.filter(user=request.user).delete()
            return JsonResponse({'success': True, 'message': 'Biometric removed successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

@require_http_methods(["GET"])
def get_biometric_status(request):
    """Get biometric authentication status - accessible without login"""
    from accounts.models import BiometricCredential
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    username = request.GET.get('username')
    
    # Always allow public access to check status for login page
    if username:
        try:
            user = User.objects.get(username=username)
            has_biometric = BiometricCredential.objects.filter(user=user, is_active=True).exists()
            return JsonResponse({
                'has_biometric': has_biometric,
                'credential_count': 1 if has_biometric else 0
            })
        except User.DoesNotExist:
            return JsonResponse({'has_biometric': False, 'credential_count': 0})
    
    # For authenticated users
    if request.user.is_authenticated:
        credentials = BiometricCredential.objects.filter(user=request.user, is_active=True)
        return JsonResponse({
            'has_biometric': credentials.exists(),
            'credential_count': credentials.count()
        })
    
    return JsonResponse({'has_biometric': False, 'credential_count': 0})
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render

@staff_member_required
def admin_analytics(request):
    """Admin analytics dashboard"""
    from books.models import Book, Category
    from .models import BookDownload, BookView, BookReadOnline, UserDownloadLimit
    from django.db.models import Sum, Avg, Count
    from django.utils import timezone
    from datetime import timedelta
    
    today = timezone.now().date()
    last_month = today - timedelta(days=30)
    
    total_downloads = BookDownload.objects.count()
    total_views = BookView.objects.count()
    total_reads = BookReadOnline.objects.count()
    active_users = UserDownloadLimit.objects.filter(downloads_today__gt=0).count()
    
    recent_downloads_count = BookDownload.objects.filter(downloaded_at__date__gte=last_month).count()
    
    # Popular books
    popular_books = Book.objects.filter(is_active=True).annotate(
        total_downloads=Sum('downloads_count')
    ).order_by('-total_downloads')[:10]
    
    context = {
        'total_downloads': total_downloads,
        'total_views': total_views,
        'total_reads': total_reads,
        'active_users': active_users,
        'recent_downloads_count': recent_downloads_count,
        'popular_books': popular_books,
    }
    
    return render(request, 'downloads/admin_dashboard.html', context)
@csrf_exempt
def biometric_login(request):
    """Handle biometric login via AJAX - Simplified version"""
    print("=" * 50)
    print("BIOMETRIC LOGIN ATTEMPT")
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
    try:
        # Parse request body
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            username = data.get('username')
        else:
            username = request.POST.get('username')
        
        print(f"Username: {username}")
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        # Find user
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        try:
            user = User.objects.get(username=username)
            print(f"User found: {user.username}")
        except User.DoesNotExist:
            try:
                user = User.objects.get(email=username)
                print(f"User found by email: {user.email}")
            except User.DoesNotExist:
                print(f"User not found: {username}")
                return JsonResponse({'error': 'User not found'}, status=404)
        
        # Check if user has biometric
        from accounts.models import BiometricCredential
        has_biometric = BiometricCredential.objects.filter(user=user, is_active=True).exists()
        print(f"Has biometric: {has_biometric}")
        
        if not has_biometric:
            return JsonResponse({
                'success': False,
                'error': 'Biometric not set up. Please set up in Security Settings first.'
            }, status=401)
        
        # Login the user
        from django.contrib.auth import login
        login(request, user)
        print(f"User logged in successfully: {user.username}")
        
        return JsonResponse({
            'success': True,
            'redirect_url': '/',
            'message': 'Login successful'
        })
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
@login_required
def enable_2fa(request):
    """Enable two-factor authentication"""
    if request.method == 'POST':
        try:
            import pyotp
            data = json.loads(request.body)
            otp_code = data.get('otp_code')
            secret_key = request.session.get('2fa_secret_key')
            
            if not secret_key:
                return JsonResponse({'success': False, 'error': 'Session expired. Please try again.'}, status=400)
            
            totp = pyotp.TOTP(secret_key)
            if totp.verify(otp_code):
                from accounts.models import TwoFactorAuth
                
                TwoFactorAuth.objects.update_or_create(
                    user=request.user,
                    defaults={
                        'secret_key': secret_key,
                        'is_enabled': True
                    }
                )
                request.session.pop('2fa_secret_key', None)
                return JsonResponse({'success': True, 'message': '2FA enabled successfully'})
            else:
                return JsonResponse({'success': False, 'error': 'Invalid OTP code'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    # GET request - generate secret key
    import pyotp
    secret_key = pyotp.random_base32()
    request.session['2fa_secret_key'] = secret_key
    
    totp = pyotp.TOTP(secret_key)
    provisioning_uri = totp.provisioning_uri(request.user.email, issuer_name="Bantu Books Zambia")
    
    # Generate QR code
    import qrcode
    from io import BytesIO
    import base64
    
    qr = qrcode.QRCode(box_size=10, border=4)
    qr.add_data(provisioning_uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    return JsonResponse({
        'success': True,
        'secret_key': secret_key,
        'qr_code': qr_code_base64,
        'provisioning_uri': provisioning_uri
    })


@login_required
def disable_2fa(request):
    """Disable two-factor authentication"""
    if request.method == 'POST':
        try:
            from accounts.models import TwoFactorAuth
            
            TwoFactorAuth.objects.filter(user=request.user).delete()
            return JsonResponse({'success': True, 'message': '2FA disabled successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@login_required
def qrcode_view(request):
    """Generate QR code for 2FA setup"""
    try:
        import pyotp
        import qrcode
        from io import BytesIO
        import base64
        
        secret_key = request.session.get('2fa_secret_key')
        if not secret_key:
            return JsonResponse({'success': False, 'error': 'No 2FA session found'}, status=400)
        
        totp = pyotp.TOTP(secret_key)
        provisioning_uri = totp.provisioning_uri(request.user.email, issuer_name="Bantu Books Zambia")
        
        qr = qrcode.QRCode(box_size=10, border=4)
        qr.add_data(provisioning_uri)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        return JsonResponse({
            'success': True,
            'qr_code': qr_code_base64,
            'secret_key': secret_key
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def generate_backup_codes(request):
    """Generate backup codes for 2FA"""
    if request.method == 'POST':
        try:
            from accounts.models import TwoFactorAuth
            import secrets
            
            # Generate 10 backup codes
            backup_codes = [secrets.token_hex(5) for _ in range(10)]
            
            twofa, created = TwoFactorAuth.objects.get_or_create(user=request.user)
            twofa.backup_codes = backup_codes
            twofa.save()
            
            return JsonResponse({
                'success': True,
                'backup_codes': backup_codes,
                'message': 'Backup codes generated successfully. Save them in a safe place.'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth import get_user_model

User = get_user_model()

def forgot_password(request):
    """Request password reset email"""
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            # Generate token
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            # Build reset link
            reset_link = request.build_absolute_uri(
                reverse('accounts:reset_password', kwargs={'uidb64': uid, 'token': token})
            )
            
            # Send email
            send_mail(
                'Password Reset Request',
                f'Click the link to reset your password: {reset_link}',
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            messages.success(request, 'Password reset link sent to your email.')
            return redirect('accounts:login')
        except User.DoesNotExist:
            messages.error(request, 'No user found with this email address.')
    
    return render(request, 'accounts/forgot_password.html')


def reset_password(request, uidb64, token):
    """Reset password page"""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    if user and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                messages.success(request, 'Password reset successfully. Please login.')
                return redirect('accounts:login')
            else:
                messages.error(request, 'Passwords do not match.')
        
        return render(request, 'accounts/reset_password.html', {'validlink': True})
    else:
        return render(request, 'accounts/reset_password.html', {'validlink': False})

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, Count
from datetime import datetime, timedelta
from downloads.models import BookDownload
from downloads.models import UserDownloadLimit
from django.contrib import messages

@login_required
def subscription_dashboard(request):
    """Display user's subscription information and usage statistics"""
    
    # Get or create user download limit
    download_limit, created = UserDownloadLimit.objects.get_or_create(
        user=request.user,
        defaults={
            'subscription_tier': request.user.user_type if request.user.user_type in ['basic', 'premium', 'lifetime'] else 'basic'
        }
    )
    
    # Update subscription tier based on user type
    user_type_map = {
        'basic': 'basic',
        'premium': 'premium',
        'lifetime': 'lifetime'
    }
    
    current_tier = user_type_map.get(request.user.user_type, 'basic')
    if download_limit.subscription_tier != current_tier:
        download_limit.subscription_tier = current_tier
        limits = download_limit.get_limit_by_tier()
        download_limit.daily_limit = limits.get('daily', 20) if limits.get('daily') else None
        download_limit.monthly_limit = limits.get('monthly', 500) if limits.get('monthly') else None
        download_limit.save()
    
    # Get download statistics
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)
    
    # Today's downloads
    today_downloads = BookDownload.objects.filter(
        user=request.user,
        downloaded_at__date=today
    ).count()
    
    # This week's downloads
    week_downloads = BookDownload.objects.filter(
        user=request.user,
        downloaded_at__date__gte=start_of_week
    ).count()
    
    # This month's downloads
    month_downloads = BookDownload.objects.filter(
        user=request.user,
        downloaded_at__date__gte=start_of_month
    ).count()
    
    # Total downloads all time
    total_downloads = BookDownload.objects.filter(
        user=request.user
    ).count()
    
    # Update download limit counts
    download_limit.downloads_today = today_downloads
    download_limit.downloads_this_month = month_downloads
    download_limit.downloads_total = total_downloads
    download_limit.save()
    
    # Get recent downloads (last 10)
    recent_downloads = BookDownload.objects.filter(
        user=request.user
    ).select_related('book').order_by('-downloaded_at')[:10]
    
    # Calculate remaining counts
    if download_limit.subscription_tier in ['premium', 'lifetime']:
        remaining_today = "Unlimited"
        remaining_month = "Unlimited"
    else:
        remaining_today = max(0, (download_limit.daily_limit or 0) - today_downloads)
        remaining_month = max(0, (download_limit.monthly_limit or 0) - month_downloads)
    
    # Check if subscription is active
    is_subscription_active = False
    subscription_expiry = None
    
    if request.user.user_type == 'premium':
        if hasattr(request.user, 'subscription_expiry') and request.user.subscription_expiry:
            if request.user.subscription_expiry > timezone.now():
                is_subscription_active = True
                subscription_expiry = request.user.subscription_expiry
            else:
                # Subscription expired
                request.user.user_type = 'basic'
                request.user.save()
                messages.warning(request, 'Your premium subscription has expired. Please renew to continue enjoying premium benefits.')
        else:
            is_subscription_active = True  # Legacy premium users
    elif request.user.user_type == 'lifetime':
        is_subscription_active = True
    
    # Prepare plan information
    plan_info = {
        'basic': {
            'name': 'Basic (Mumbi)',
            'price': 'Free',
            'download_limit': '20 downloads/day',
            'features': [
                '20 downloads per day',
                'Access to free books only',
                'Basic support',
                'Standard download speed'
            ]
        },
        'premium': {
            'name': 'Premium (Chibombo)',
            'price': 'K35/month',
            'download_limit': 'Unlimited',
            'features': [
                'Unlimited downloads',
                'Access to ALL books',
                'No advertisements',
                'Priority support',
                'Early access to new books',
                'Download up to 10,000 books/month'
            ]
        },
        'lifetime': {
            'name': 'Lifetime (Chifumu)',
            'price': 'K800 one-time',
            'download_limit': 'Unlimited',
            'features': [
                'Unlimited downloads forever',
                'Access to ALL books',
                'No advertisements',
                'Priority support',
                'Early access to new books',
                'Name in supporters list',
                'Never expires'
            ]
        }
    }
    
    current_plan = plan_info.get(current_tier, plan_info['basic'])
    
    # Get monthly download data for chart
    monthly_data = []
    for i in range(6):
        month_date = start_of_month - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        
        month_download_count = BookDownload.objects.filter(
            user=request.user,
            downloaded_at__date__gte=month_start,
            downloaded_at__date__lte=month_end
        ).count()
        
        monthly_data.append({
            'month': month_start.strftime('%B %Y'),
            'downloads': month_download_count
        })
    
    context = {
        'current_plan': current_plan,
        'current_tier': current_tier,
        'total_downloads': total_downloads,
        'today_downloads': today_downloads,
        'week_downloads': week_downloads,
        'month_downloads': month_downloads,
        'remaining_today': remaining_today,
        'remaining_month': remaining_month,
        'daily_limit': download_limit.daily_limit,
        'monthly_limit': download_limit.monthly_limit,
        'recent_downloads': recent_downloads,
        'is_subscription_active': is_subscription_active,
        'subscription_expiry': subscription_expiry,
        'monthly_data': monthly_data,
        'plan_info': plan_info,
    }
    
    return render(request, 'accounts/subscription_dashboard.html', context)




from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta
from accounts.models import User
from downloads.models import Transaction, UserDownloadLimit, BookDownload
from books.models import Book, Category, GradeLevel, Language, BookRequest
from books.models import ContributorApplication, ContributorWithdrawal

@staff_member_required
def admin_subscription_dashboard(request):
    """Admin dashboard for managing subscriptions, payments, and users"""
    
    # ============ STATISTICS ============
    # User statistics
    total_users = User.objects.count()
    premium_users = User.objects.filter(user_type='premium').count()
    lifetime_users = User.objects.filter(user_type='lifetime').count()
    basic_users = User.objects.filter(user_type='basic').count()
    new_users_today = User.objects.filter(date_joined__date=timezone.now().date()).count()
    new_users_this_week = User.objects.filter(date_joined__date__gte=timezone.now().date() - timedelta(days=7)).count()
    
    # Payment statistics
    total_revenue = Transaction.objects.filter(status='completed').aggregate(total=Sum('amount'))['total'] or 0
    pending_payments = Transaction.objects.filter(status='pending').count()
    completed_payments = Transaction.objects.filter(status='completed').count()
    failed_payments = Transaction.objects.filter(status='failed').count()
    
    # Monthly revenue for chart
    monthly_revenue = []
    for i in range(6):
        month_date = timezone.now().date().replace(day=1) - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if i == 0:
            month_end = timezone.now().date()
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        
        revenue = Transaction.objects.filter(
            status='completed',
            payment_confirmed_at__date__gte=month_start,
            payment_confirmed_at__date__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        monthly_revenue.append({
            'month': month_start.strftime('%B %Y'),
            'revenue': float(revenue)
        })
    
    # Download statistics
    total_downloads = BookDownload.objects.count()
    downloads_today = BookDownload.objects.filter(downloaded_at__date=timezone.now().date()).count()
    downloads_this_week = BookDownload.objects.filter(downloaded_at__date__gte=timezone.now().date() - timedelta(days=7)).count()
    
    # Popular books
    popular_books = Book.objects.annotate(
        download_count=Count('downloads')
    ).order_by('-download_count')[:5]
    
    # ============ RECENT TRANSACTIONS ============
    recent_transactions = Transaction.objects.select_related('user').order_by('-created_at')[:10]
    
    # ============ PENDING PAYMENTS ============
    pending_transactions = Transaction.objects.filter(
        status='pending'
    ).select_related('user').order_by('-created_at')
    
    # ============ RECENT USERS ============
    recent_users = User.objects.order_by('-date_joined')[:10]
    
    # ============ PLAN DISTRIBUTION ============
    plan_distribution = {
        'basic': basic_users,
        'premium': premium_users,
        'lifetime': lifetime_users,
    }
    
    # ============ SUBSCRIPTION EXPIRING SOON ============
    expiring_soon = User.objects.filter(
        user_type='premium',
        subscription_expiry__isnull=False,
        subscription_expiry__lte=timezone.now() + timedelta(days=7),
        subscription_expiry__gte=timezone.now()
    ).count()
    
    context = {
        # Statistics
        'total_users': total_users,
        'premium_users': premium_users,
        'lifetime_users': lifetime_users,
        'basic_users': basic_users,
        'new_users_today': new_users_today,
        'new_users_this_week': new_users_this_week,
        'total_revenue': total_revenue,
        'pending_payments': pending_payments,
        'completed_payments': completed_payments,
        'failed_payments': failed_payments,
        'total_downloads': total_downloads,
        'downloads_today': downloads_today,
        'downloads_this_week': downloads_this_week,
        'expiring_soon': expiring_soon,
        
        # Data tables
        'recent_transactions': recent_transactions,
        'pending_transactions': pending_transactions,
        'recent_users': recent_users,
        'popular_books': popular_books,
        
        # Charts
        'monthly_revenue': monthly_revenue,
        'plan_distribution': plan_distribution,
    }
    
    return render(request, 'admin/subscription_dashboard.html', context)

@staff_member_required
def admin_subscription_payments(request):
    """View all payment transactions"""
    from downloads.models import Transaction
    
    status_filter = request.GET.get('status', '')
    plan_filter = request.GET.get('plan', '')
    
    transactions = Transaction.objects.select_related('user').order_by('-created_at')
    
    if status_filter:
        transactions = transactions.filter(status=status_filter)
    if plan_filter:
        transactions = transactions.filter(plan_type=plan_filter)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(transactions, 50)
    page = request.GET.get('page', 1)
    transactions_page = paginator.get_page(page)
    
    # Create choices for filters
    status_choices = [
        ('pending', 'Pending'),
        ('processing', 'Processing'),
        ('completed', 'Completed'),
        ('failed', 'Failed'),
        ('cancelled', 'Cancelled'),
        ('refunded', 'Refunded'),
    ]
    
    plan_choices = [
        ('monthly', 'Monthly Premium'),
        ('quarterly', 'Quarterly Premium'),
        ('yearly', 'Yearly Premium'),
        ('lifetime', 'Lifetime Access'),
        ('book_purchase', 'Book Purchase'),
    ]
    
  
    total_amount = transactions.aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'transactions': transactions_page,
        'status_filter': status_filter,
        'plan_filter': plan_filter,
        'total_amount': total_amount,
        'status_choices': status_choices,
        'plan_choices': plan_choices,
    }
    
    return render(request, 'admin/payments.html', context)


@staff_member_required
def admin_update_payment_status(request, transaction_id):
    """Update payment status via AJAX"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            new_status = data.get('status')
            
            transaction = Transaction.objects.get(transaction_id=transaction_id)
            old_status = transaction.status
            transaction.status = new_status
            
            if new_status == 'completed' and not transaction.payment_confirmed_at:
                transaction.payment_confirmed_at = timezone.now()
            
            transaction.save()
            
            # If payment completed, ensure user is upgraded
            if new_status == 'completed' and old_status != 'completed':
                from accounts.views import upgrade_user_subscription
                upgrade_user_subscription(transaction.user, transaction.plan_type, transaction)
            
            return JsonResponse({'success': True, 'message': 'Payment status updated'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@staff_member_required
def admin_subscription_plans(request):
    """Manage subscription plans"""
    from accounts.views import subscription_plans
    
    if request.method == 'POST':
        # Update plan settings
        import json
        data = json.loads(request.body)
        # Save plan settings to database or config
        messages.success(request, 'Plans updated successfully')
        return JsonResponse({'success': True})
    
    context = {
        'plans': subscription_plans,
    }
    return render(request, 'admin/plans.html', context)


@staff_member_required
def admin_cancel_subscription(request, user_id):
    """Cancel a user's subscription"""
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            user.user_type = 'basic'
            user.subscription_expiry = None
            user.save()
            
            # Update download limits
            from downloads.models import UserDownloadLimit
            limit, _ = UserDownloadLimit.objects.get_or_create(user=user)
            limit.subscription_tier = 'basic'
            limit.daily_limit = 20
            limit.monthly_limit = 500
            limit.save()
            
            messages.success(request, f'Subscription cancelled for {user.email}')
            return redirect('accounts:admin_subscription_dashboard')
        except User.DoesNotExist:
            messages.error(request, 'User not found')
            return redirect('accounts:admin_subscription_dashboard')
    
    return redirect('accounts:admin_subscription_dashboard')


import csv
import json
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO

@staff_member_required
def export_subscription_data(request):
    """Export subscription data to CSV, Excel, or PDF"""
    
    # Get format from query parameter (csv, excel, pdf)
    export_format = request.GET.get('format', 'csv')
    
    # Get date range filter
    date_range = request.GET.get('date_range', 'all')
    
    # Build query based on date range
    from downloads.models import Transaction, BookDownload
    from accounts.models import User
    from django.db.models import Sum, Count
    
    transactions = Transaction.objects.select_related('user').all()
    users = User.objects.all()
    downloads = BookDownload.objects.select_related('book', 'user').all()
    
    now = timezone.now()
    
    if date_range == 'today':
        transactions = transactions.filter(created_at__date=now.date())
        users = users.filter(date_joined__date=now.date())
        downloads = downloads.filter(downloaded_at__date=now.date())
    elif date_range == 'week':
        week_ago = now - timedelta(days=7)
        transactions = transactions.filter(created_at__gte=week_ago)
        users = users.filter(date_joined__gte=week_ago)
        downloads = downloads.filter(downloaded_at__gte=week_ago)
    elif date_range == 'month':
        month_ago = now - timedelta(days=30)
        transactions = transactions.filter(created_at__gte=month_ago)
        users = users.filter(date_joined__gte=month_ago)
        downloads = downloads.filter(downloaded_at__gte=month_ago)
    elif date_range == 'year':
        year_ago = now - timedelta(days=365)
        transactions = transactions.filter(created_at__gte=year_ago)
        users = users.filter(date_joined__gte=year_ago)
        downloads = downloads.filter(downloaded_at__gte=year_ago)
    
    # Calculate statistics - FIX: Use .get() or check if key exists
    total_revenue_result = transactions.filter(status='completed').aggregate(total=Sum('amount'))
    total_revenue = total_revenue_result.get('total') or 0
    
    total_transactions = transactions.count()
    completed_transactions = transactions.filter(status='completed').count()
    pending_transactions = transactions.filter(status='pending').count()
    failed_transactions = transactions.filter(status='failed').count()
    
    total_users_count = users.count()
    premium_users = users.filter(user_type='premium').count()
    lifetime_users = users.filter(user_type='lifetime').count()
    basic_users = users.filter(user_type='basic').count()
    
    total_downloads_count = downloads.count()
    
    if export_format == 'excel':
        return export_to_excel(transactions, users, downloads, total_revenue, total_transactions, completed_transactions, pending_transactions, failed_transactions, total_users_count, premium_users, lifetime_users, basic_users, total_downloads_count, date_range)
    elif export_format == 'pdf':
        return export_to_pdf(transactions, users, downloads, total_revenue, total_transactions, completed_transactions, pending_transactions, failed_transactions, total_users_count, premium_users, lifetime_users, basic_users, total_downloads_count, date_range)
    else:
        return export_to_csv(transactions, users, downloads, total_revenue, total_transactions, completed_transactions, pending_transactions, failed_transactions, total_users_count, premium_users, lifetime_users, basic_users, total_downloads_count, date_range)


def export_to_csv(transactions, users, downloads, total_revenue, total_transactions, completed_transactions, pending_transactions, failed_transactions, total_users_count, premium_users, lifetime_users, basic_users, total_downloads_count, date_range):
    """Export data to CSV format"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="subscription_data_{date_range}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow(['BANTU BOOKS ZAMBIA - SUBSCRIPTION DATA EXPORT'])
    writer.writerow([f'Export Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([f'Date Range: {date_range.upper()}'])
    writer.writerow([])
    
    # Summary Statistics
    writer.writerow(['SUMMARY STATISTICS'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Total Revenue', f'K{total_revenue:,.2f}'])
    writer.writerow(['Total Transactions', total_transactions])
    writer.writerow(['Completed Transactions', completed_transactions])
    writer.writerow(['Pending Transactions', pending_transactions])
    writer.writerow(['Failed Transactions', failed_transactions])
    writer.writerow(['Total Users', total_users_count])
    writer.writerow(['Premium Users', premium_users])
    writer.writerow(['Lifetime Users', lifetime_users])
    writer.writerow(['Basic Users', basic_users])
    writer.writerow(['Total Downloads', total_downloads_count])
    writer.writerow([])
    
    # Transactions Detail
    writer.writerow(['TRANSACTION DETAILS'])
    writer.writerow(['Transaction ID', 'User Email', 'User Type', 'Amount', 'Plan Type', 'Payment Method', 'Status', 'Created At', 'Completed At'])
    
    for transaction in transactions[:1000]:  # Limit to 1000 for performance
        writer.writerow([
            transaction.transaction_id,
            transaction.user.email if transaction.user else 'N/A',
            transaction.user.user_type if transaction.user else 'N/A',
            f'K{transaction.amount}',
            transaction.plan_type or 'N/A',
            transaction.payment_method or 'N/A',
            transaction.status,
            transaction.created_at.strftime('%Y-%m-%d %H:%M:%S') if transaction.created_at else 'N/A',
            transaction.payment_confirmed_at.strftime('%Y-%m-%d %H:%M:%S') if transaction.payment_confirmed_at else 'N/A'
        ])
    
    writer.writerow([])
    
    # Users Detail
    writer.writerow(['USER DETAILS'])
    writer.writerow(['Email', 'Username', 'User Type', 'Date Joined', 'Last Login', 'Subscription Expiry'])
    
    for user in users[:500]:  # Limit to 500 for performance
        writer.writerow([
            user.email,
            user.username,
            user.user_type,
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else 'N/A',
            user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'N/A',
            user.subscription_expiry.strftime('%Y-%m-%d') if hasattr(user, 'subscription_expiry') and user.subscription_expiry else 'N/A'
        ])
    
    writer.writerow([])
    
    # Downloads Detail
    writer.writerow(['DOWNLOAD DETAILS'])
    writer.writerow(['Book Title', 'Book Author', 'User Email', 'Downloaded At', 'IP Address'])
    
    for download in downloads[:1000]:  # Limit to 1000 for performance
        writer.writerow([
            download.book.title if download.book else 'N/A',
            download.book.author if download.book else 'N/A',
            download.user.email if download.user else 'Anonymous',
            download.downloaded_at.strftime('%Y-%m-%d %H:%M:%S') if download.downloaded_at else 'N/A',
            download.ip_address or 'N/A'
        ])
    
    return response


def export_to_excel(transactions, users, downloads, total_revenue, total_transactions, completed_transactions, pending_transactions, failed_transactions, total_users_count, premium_users, lifetime_users, basic_users, total_downloads_count, date_range):
    """Export data to Excel format using openpyxl"""
    
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    title_font = Font(bold=True, size=16)
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws_summary['A1'] = "BANTU BOOKS ZAMBIA - SUBSCRIPTION DATA EXPORT"
    ws_summary['A1'].font = title_font
    ws_summary['A2'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A3'] = f"Date Range: {date_range.upper()}"
    
    # Summary Statistics
    ws_summary['A5'] = "SUMMARY STATISTICS"
    ws_summary['A5'].font = Font(bold=True, size=12)
    
    summary_data = [
        ("Total Revenue", f"K{total_revenue:,.2f}"),
        ("Total Transactions", total_transactions),
        ("Completed Transactions", completed_transactions),
        ("Pending Transactions", pending_transactions),
        ("Failed Transactions", failed_transactions),
        ("Total Users", total_users_count),
        ("Premium Users", premium_users),
        ("Lifetime Users", lifetime_users),
        ("Basic Users", basic_users),
        ("Total Downloads", total_downloads_count),
    ]
    
    for i, (key, value) in enumerate(summary_data, start=7):
        ws_summary[f'A{i}'] = key
        ws_summary[f'B{i}'] = value
        ws_summary[f'A{i}'].font = Font(bold=True)
    
    # Transactions Sheet
    ws_transactions = wb.create_sheet("Transactions")
    
    # Headers
    headers = ['Transaction ID', 'User Email', 'User Type', 'Amount', 'Plan Type', 'Payment Method', 'Status', 'Created At', 'Completed At']
    for col, header in enumerate(headers, 1):
        cell = ws_transactions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = cell_border
    
    # Data
    for row, transaction in enumerate(transactions, 2):
        ws_transactions.cell(row=row, column=1, value=transaction.transaction_id)
        ws_transactions.cell(row=row, column=2, value=transaction.user.email)
        ws_transactions.cell(row=row, column=3, value=transaction.user.user_type)
        ws_transactions.cell(row=row, column=4, value=f"K{transaction.amount}")
        ws_transactions.cell(row=row, column=5, value=transaction.plan_type or 'N/A')
        ws_transactions.cell(row=row, column=6, value=transaction.payment_method or 'N/A')
        ws_transactions.cell(row=row, column=7, value=transaction.status)
        ws_transactions.cell(row=row, column=8, value=transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws_transactions.cell(row=row, column=9, value=transaction.payment_confirmed_at.strftime('%Y-%m-%d %H:%M:%S') if transaction.payment_confirmed_at else 'N/A')
    
    # Auto-adjust column widths
    for col in ws_transactions.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_transactions.column_dimensions[column].width = adjusted_width
    
    # Users Sheet
    ws_users = wb.create_sheet("Users")
    
    user_headers = ['Email', 'Username', 'User Type', 'Date Joined', 'Last Login', 'Subscription Expiry']
    for col, header in enumerate(user_headers, 1):
        cell = ws_users.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = cell_border
    
    for row, user in enumerate(users, 2):
        ws_users.cell(row=row, column=1, value=user.email)
        ws_users.cell(row=row, column=2, value=user.username)
        ws_users.cell(row=row, column=3, value=user.user_type)
        ws_users.cell(row=row, column=4, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else 'N/A')
        ws_users.cell(row=row, column=5, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'N/A')
        ws_users.cell(row=row, column=6, value=user.subscription_expiry.strftime('%Y-%m-%d') if hasattr(user, 'subscription_expiry') and user.subscription_expiry else 'N/A')
    
    # Downloads Sheet
    ws_downloads = wb.create_sheet("Downloads")
    
    download_headers = ['Book Title', 'Book Author', 'User Email', 'Downloaded At', 'IP Address']
    for col, header in enumerate(download_headers, 1):
        cell = ws_downloads.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = cell_border
    
    for row, download in enumerate(downloads[:5000], 2):  # Limit to 5000 for performance
        ws_downloads.cell(row=row, column=1, value=download.book.title)
        ws_downloads.cell(row=row, column=2, value=download.book.author)
        ws_downloads.cell(row=row, column=3, value=download.user.email if download.user else 'Anonymous')
        ws_downloads.cell(row=row, column=4, value=download.downloaded_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws_downloads.cell(row=row, column=5, value=download.ip_address or 'N/A')
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="subscription_data_{date_range}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


def export_to_pdf(transactions, users, downloads, total_revenue, total_transactions, completed_transactions, pending_transactions, failed_transactions, total_users_count, premium_users, lifetime_users, basic_users, total_downloads_count, date_range):
    """Export data to PDF format using reportlab"""
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []
    
    # Title Style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#059669'),
        alignment=1,
        spaceAfter=30
    )
    
    # Add Title
    elements.append(Paragraph("Bantu Books Zambia - Subscription Data Export", title_style))
    elements.append(Paragraph(f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Paragraph(f"Date Range: {date_range.upper()}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary Statistics Table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Revenue', f'K{total_revenue:,.2f}'],
        ['Total Transactions', str(total_transactions)],
        ['Completed Transactions', str(completed_transactions)],
        ['Pending Transactions', str(pending_transactions)],
        ['Failed Transactions', str(failed_transactions)],
        ['Total Users', str(total_users_count)],
        ['Premium Users', str(premium_users)],
        ['Lifetime Users', str(lifetime_users)],
        ['Basic Users', str(basic_users)],
        ['Total Downloads', str(total_downloads_count)],
    ]
    
    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Recent Transactions Table
    elements.append(Paragraph("Recent Transactions", styles['Heading2']))
    elements.append(Spacer(1, 10))
    
    trans_data = [['Transaction ID', 'User', 'Amount', 'Plan', 'Status', 'Date']]
    for transaction in transactions[:20]:  # Limit to 20 transactions for PDF
        trans_data.append([
            transaction.transaction_id[:12],
            transaction.user.email[:20],
            f'K{transaction.amount}',
            transaction.plan_type or 'N/A',
            transaction.status,
            transaction.created_at.strftime('%Y-%m-%d')
        ])
    
    trans_table = Table(trans_data, colWidths=[80, 100, 60, 60, 80, 80])
    trans_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(trans_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="subscription_data_{date_range}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    return response